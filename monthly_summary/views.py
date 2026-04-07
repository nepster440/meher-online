from django.views.generic import TemplateView
from data_entry.models import DailyEntry
from datetime import datetime
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth.mixins import LoginRequiredMixin

# Create your views here.
class MonthlySummaryView(LoginRequiredMixin, TemplateView):
    template_name = "monthly_summary.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        request = self.request

        month = int(request.GET.get("month", datetime.today().month))
        year = int(request.GET.get("year", datetime.today().year))

        entries = DailyEntry.objects.filter(
            date__month=month,
            date__year=year
        ).order_by("date")

        # 🔥 GROUP BY DATE
        grouped = defaultdict(lambda: {"cash": 0, "gpay": 0})

        for e in entries:
            cash = e.xe + e.press + e.color
            gpay = e.online + e.xg + e.pg + e.og + e.cg

            grouped[e.date]["cash"] += cash
            grouped[e.date]["gpay"] += gpay

        data = []
        total_cash = 0
        total_gpay = 0

        for i, (date, val) in enumerate(sorted(grouped.items()), start=1):

            total_cash += val["cash"]
            total_gpay += val["gpay"]

            data.append({
                "no": i,
                "date": date.strftime("%d-%m-%Y"),
                "cash": val["cash"],
                "gpay": val["gpay"],
                "total": val["cash"] + val["gpay"]
            })

        context["data"] = data
        context["cash_total"] = total_cash
        context["gpay_total"] = total_gpay
        context["final_total"] = total_cash + total_gpay

        context["selected_month"] = month
        context["selected_year"] = year

        return context

# from openpyxl.utils import get_column_letter

def auto_width(ws):
    column_widths = {}

    for row in ws.iter_rows():
        for cell in row:

            # Skip merged cells safely
            if cell.coordinate in ws.merged_cells:
                continue

            try:
                if cell.value:
                    col_index = cell.column
                    col_letter = get_column_letter(col_index)

                    length = len(str(cell.value))

                    if col_letter not in column_widths:
                        column_widths[col_letter] = length
                    else:
                        column_widths[col_letter] = max(column_widths[col_letter], length)

            except:
                pass

    # Apply width
    for col_letter, max_len in column_widths.items():
        adjusted_width = (max_len + 4) * 1.2   # tuned for Montserrat
        ws.column_dimensions[col_letter].width = max(adjusted_width, 12)
        ws.row_dimensions[4].height = 28


# 🔥 EXPORT EXCEL
def export_excel(request):

    month = int(request.GET.get("month"))
    year = int(request.GET.get("year"))

    wb = Workbook()
    wb.remove(wb.active)

    entries = DailyEntry.objects.filter(
        date__month=month,
        date__year=year
    ).order_by("date")

    grouped = {}

    for e in entries:
        date_str = e.date.strftime("%d-%m-%Y")
        grouped.setdefault(date_str, []).append(e)

    # ===== STYLES =====
    title_font = Font(name="Montserrat", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Montserrat", size=14, bold=True, color="FFFFFF")
    normal_font = Font(name="Montserrat", size=14, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(start_color="6E0C9C", end_color="6E0C9C", fill_type="solid")
    data_fill = PatternFill(start_color="2A003F", end_color="2A003F", fill_type="solid")
    total_fill = PatternFill(start_color="DE7300", end_color="DE7300", fill_type="solid")

    border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF")
    )

    total_cash = 0
    total_gpay = 0

    # ===== DAILY SHEETS =====
    for date, rows in grouped.items():

        ws = wb.create_sheet(title=date)

        # ===== BRAND HEADER =====
        ws.merge_cells("A1:J1")
        ws["A1"] = "MEHER COMPUTING AND ONLINE"
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws["A1"].fill = header_fill

        # ===== DATE + MONTH HEADER =====
        ws.merge_cells("A2:J2")
        ws["A2"] = f"Date: {date} | Month: {month}-{year}"
        ws["A2"].font = header_font
        ws["A2"].alignment = center
        ws["A2"].fill = data_fill

        # ===== TABLE HEADER =====
        headers = ["No", "XE", "PRESS", "ONLINE", "COLOR", "XG", "PG", "OG", "CG", "Total"]

        ws.append([])
        ws.append(headers)

        for col in range(1, 11):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # ===== DATA =====
        day_total = 0

        for i, r in enumerate(rows, start=1):

            total = r.xe + r.press + r.online + r.color + r.xg + r.pg + r.og + r.cg

            ws.append([
                i, r.xe, r.press, r.online, r.color,
                r.xg, r.pg, r.og, r.cg, total
            ])

            current_row = ws.max_row

            for col in range(1, 11):
                cell = ws.cell(row=current_row, column=col)
                cell.font = normal_font
                cell.fill = data_fill
                cell.alignment = center
                cell.border = border

            day_total += total
            total_cash += r.xe + r.press + r.color
            total_gpay += r.online + r.xg + r.pg + r.og + r.cg

        # ===== GRAND TOTAL =====
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "", "GRAND TOTAL", day_total])

        last_row = ws.max_row

        for col in range(1, 11):
            cell = ws.cell(row=last_row, column=col)
            cell.font = header_font
            cell.fill = total_fill
            cell.alignment = center
            cell.border = border

        # AUTO WIDTH
        auto_width(ws)

    # ===== MONTHLY SUMMARY SHEET =====
    ws = wb.create_sheet(title="Monthly Summary")

    # BRAND
    ws.merge_cells("A1:D1")
    ws["A1"] = "MEHER COMPUTING AND ONLINE"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = header_fill

    # MONTH HEADER
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Monthly Summary - {month}-{year}"
    ws["A2"].font = header_font
    ws["A2"].alignment = center
    ws["A2"].fill = data_fill

    ws.append([])
    ws.append(["Date", "CASH", "GPAY", "TOTAL"])

    for col in range(1, 5):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # DATA
    for date, rows in grouped.items():

        cash = sum(r.xe + r.press + r.color for r in rows)
        gpay = sum(r.online + r.xg + r.pg + r.og + r.cg for r in rows)

        ws.append([date, cash, gpay, cash + gpay])

        current_row = ws.max_row

        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = normal_font
            cell.fill = data_fill
            cell.alignment = center
            cell.border = border

    # GRAND TOTAL
    ws.append([])
    ws.append(["GRAND TOTAL", total_cash, total_gpay, total_cash + total_gpay])

    last_row = ws.max_row

    for col in range(1, 5):
        cell = ws.cell(row=last_row, column=col)
        cell.font = header_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = border

    # AUTO WIDTH
    auto_width(ws)

    # DOWNLOAD
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Meher_Report_{month}_{year}.xlsx'

    wb.save(response)

    return response